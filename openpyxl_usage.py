import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


data = {"1班": [["01", "张三", "8"], ["02", "李四", "8"]],
        "2班": [["01", "王二", "7"], ["02", "麻子", "9"]]}

# 添加表头
header = "班级 学号 姓名 年龄"
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(header.split(" "))

# 写入数据
row_idx = 2
for class_num, res in data.items():
    sheet.merge_cells(start_row=row_idx, end_row=(row_idx + len(res) - 1), start_column=1, end_column=1)
    sheet.cell(row=row_idx, column=1).value = class_num
    for i in range(len(res)):
        sheet.cell(row=(row_idx + i), column=2).value = res[i][0]
        sheet.cell(row=(row_idx + i), column=3).value = res[i][1]
        sheet.cell(row=(row_idx + i), column=4).value = res[i][2]
    row_idx += len(res)

# 设置cell自动换行
for row in sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")

# 设置列宽
for i in range(1, sheet.max_column + 1):
    sheet.column_dimensions[get_column_letter(i)].width = 20

# 保存文件
wb.save("class_info.xlsx")
